from __future__ import annotations

import json
import warnings
from pathlib import Path
from typing import Dict, Any

import pytest
from openpyxl import load_workbook

from src.spreadsheet_builder.builder import (
    build_from_plan,
    SchemaError,
    LayoutError,
    FormulaError,
)


@pytest.fixture()
def sample_plan() -> Dict[str, Any]:
    """Return a valid build-plan matching the README example."""
    return {
        "workbook": {"filename": "innv_model.xlsx"},
        "worksheet": {
            "name": "Model",
            "columns": [
                {
                    "col": 2,
                    "header": "FY-2024",
                    "cells": [
                        {
                            "row": 2,
                            "label": "Revenue",
                            "type": "fact",
                            "unit": "dollars",
                            "value": 763900000,
                            "format": "currency_0dp",
                        },
                        {
                            "row": 3,
                            "label": "Participants",
                            "type": "fact",
                            "unit": "vanilla",
                            "value": 7020,
                            "format": "comma_0dp",
                        },
                        {
                            "row": 4,
                            "label": "Revenue per Participant",
                            "type": "calc",
                            "unit": "dollars",
                            "formula": "=B2/B3",
                            "format": "currency_2dp",
                        },
                    ],
                }
            ],
            "named_ranges": [
                {"name": "Revenue", "ref": "B2"},
                {"name": "Participants", "ref": "B3"},
            ],
        },
    }


# ---------------------------------------------------------------------------
# Happy-path test
# ---------------------------------------------------------------------------


def test_build_valid_plan(tmp_path: Path, sample_plan: Dict[str, Any]):
    output_path = build_from_plan(sample_plan, output_dir=tmp_path)
    assert output_path.exists(), "Workbook file should be written"

    wb = load_workbook(output_path, data_only=False)
    ws = wb["Model"]

    # Headers
    assert ws["B1"].value == "FY-2024"
    # Labels
    assert ws["A2"].value == "Revenue"
    assert ws["A3"].value == "Participants"
    # Values & formula
    assert ws["B2"].value == 763900000
    assert ws["B3"].value == 7020
    assert ws["B4"].value == "=B2/B3"
    # Number formats
    assert ws["B2"].number_format.startswith("$")
    assert ws["B4"].number_format.startswith("$")

    # Named ranges exist
    dn_dict = {dn.name: dn.attr_text for dn in wb.defined_names}
    assert dn_dict["Revenue"].endswith("!B2")
    assert dn_dict["Participants"].endswith("!B3")


# ---------------------------------------------------------------------------
# Filename & path validation
# ---------------------------------------------------------------------------


def test_invalid_filename_extension(sample_plan):
    sample_plan["workbook"]["filename"] = "model.xls"  # missing xlsx
    with pytest.raises(SchemaError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


def test_filename_with_path_separator(sample_plan):
    sample_plan["workbook"]["filename"] = "../model.xlsx"
    with pytest.raises(SchemaError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


# ---------------------------------------------------------------------------
# Grid layout validation
# ---------------------------------------------------------------------------


def test_column_index_less_than_two(sample_plan):
    sample_plan["worksheet"]["columns"][0]["col"] = 1
    with pytest.raises(LayoutError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


def test_row_index_less_than_two(sample_plan):
    sample_plan["worksheet"]["columns"][0]["cells"][0]["row"] = 1
    with pytest.raises(LayoutError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


def test_named_range_out_of_bounds(sample_plan):
    sample_plan["worksheet"]["named_ranges"][0]["ref"] = "Z100"
    with pytest.raises(LayoutError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


# ---------------------------------------------------------------------------
# Value & type validation
# ---------------------------------------------------------------------------


def test_value_more_than_two_decimals(sample_plan):
    sample_plan["worksheet"]["columns"][0]["cells"][0]["value"] = 123.456
    path = build_from_plan(sample_plan, output_dir=Path("/tmp"))
    wb = load_workbook(path)
    assert wb.active["B2"].value == 123.46  # rounded to 2dp


def test_unknown_unit(sample_plan):
    sample_plan["worksheet"]["columns"][0]["cells"][0]["unit"] = "euros"
    with pytest.raises(ValueError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


def test_type_fact_missing_value(sample_plan):
    # Remove 'value' field
    sample_plan["worksheet"]["columns"][0]["cells"][0].pop("value")
    with pytest.raises(SchemaError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


# ---------------------------------------------------------------------------
# Formula validation
# ---------------------------------------------------------------------------


# noqa: D401 – intentional xfail for auto-prepend behaviour
@pytest.mark.xfail(reason="Builder auto-prepends '=' instead of raising FormulaError")
def test_formula_missing_equals(sample_plan):
    cell = sample_plan["worksheet"]["columns"][0]["cells"][2]
    cell["formula"] = "B2/B3"  # missing '='
    with pytest.raises(FormulaError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


def test_formula_out_of_bounds_reference(sample_plan):
    cell = sample_plan["worksheet"]["columns"][0]["cells"][2]
    cell["formula"] = "=Z100"  # reference out of grid
    with pytest.raises(FormulaError):
        build_from_plan(sample_plan, output_dir=Path("/tmp"))


# ---------------------------------------------------------------------------
# Unit vs. format warnings
# ---------------------------------------------------------------------------


def test_percent_unit_without_percent_format_warns(sample_plan):
    # Change unit to percent but leave currency format -> should warn, not fail
    cell = sample_plan["worksheet"]["columns"][0]["cells"][0]
    cell["unit"] = "percent"
    with warnings.catch_warnings(record=True) as w:
        warnings.simplefilter("always")
        build_from_plan(sample_plan, output_dir=Path("/tmp"))
        assert any("percent" in str(warn.message).lower() for warn in w)


# ---------------------------------------------------------------------------
# Formula auto-prepend '='
# ---------------------------------------------------------------------------


def test_formula_auto_prepend(sample_plan):
    cell = sample_plan["worksheet"]["columns"][0]["cells"][2]
    cell["formula"] = "B2/B3"  # missing '='
    path = build_from_plan(sample_plan, output_dir=Path("/tmp"))
    wb = load_workbook(path, data_only=False)
    assert wb.active["B4"].value == "=B2/B3"
