from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName

__all__ = [
    "build_from_plan",
    "SchemaError",
    "LayoutError",
    "FormulaError",
]


class SchemaError(ValueError):
    """Raised when the build-plan schema is invalid or incomplete."""


class LayoutError(ValueError):
    """Raised when labels/headers/data are written outside the allowed grid."""


class FormulaError(ValueError):
    """Raised when a formula is invalid or references out-of-bounds cells."""


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


_FORMAT_MAP = {
    "comma_0dp": "#,##0",
    "comma_2dp": "#,##0.00",
    "currency_0dp": "$#,##0",
    "currency_2dp": "$#,##0.00",
    "percent_1dp": "0.0%",
    "percent_2dp": "0.00%",
}

_ALLOWED_UNITS = {"dollars", "percent", "vanilla"}
_ALLOWED_TYPES = {"fact", "assumption", "calc"}

_CELL_REF_RE = re.compile(r"\b([A-Z]+)(\d+)\b")


# ---------------------------------------------------------------------------
# Public interface
# ---------------------------------------------------------------------------

def build_from_plan(plan: Dict[str, Any], output_dir: Path = Path(".")) -> Path:  # noqa: D401,E501
    """Build an Excel workbook from a validated plan (v0.2).

    Parameters
    ----------
    plan:
        The build-plan dictionary obtained from the LLM.
    output_dir:
        Directory where the resulting ``.xlsx`` file will be written.

    Returns
    -------
    Path
        Absolute path to the saved workbook.

    Raises
    ------
    SchemaError
        If mandatory fields are missing or have the wrong type/value.
    LayoutError
        If labels, headers or data are placed outside the allowed grid.
    FormulaError
        If a formula is malformed or references cells outside the worksheet.
    ValueError
        Miscellaneous value problems (e.g. >2 decimals, unknown unit/format).
    """
    _validate_plan_root(plan)
    workbook_spec = plan["workbook"]
    worksheet_spec = plan["worksheet"]

    filename: str = workbook_spec["filename"]
    _validate_filename(filename)

    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_spec["name"]

    # Track used rows/columns to validate collisions and formula bounds
    used_labels_rows: set[int] = set()
    used_headers_cols: set[int] = set()
    max_data_row = 1
    max_data_col = 1

    # Build columns
    columns_spec: List[Dict[str, Any]] = worksheet_spec.get("columns", [])
    if not isinstance(columns_spec, list) or not columns_spec:
        raise SchemaError("'columns' must be a non-empty list")

    for col_obj in columns_spec:
        col_index = col_obj.get("col")
        header = col_obj.get("header")
        cells_spec = col_obj.get("cells")

        if not isinstance(col_index, int):
            raise SchemaError("Column index 'col' must be an integer")
        if col_index < 2:
            raise LayoutError("Data columns must start at column 2 (column B)")
        if header is None:
            raise SchemaError("Each column must include a 'header' field")
        if not isinstance(cells_spec, list) or not cells_spec:
            raise SchemaError("'cells' must be a non-empty list for each column")

        # Write header (Row 1)
        if ws.cell(row=1, column=col_index).value not in (None, ""):
            raise LayoutError(f"Header cell {get_column_letter(col_index)}1 already occupied")
        ws.cell(row=1, column=col_index, value=header)
        used_headers_cols.add(col_index)
        max_data_col = max(max_data_col, col_index)

        # Process cells
        for cell_obj in cells_spec:
            _write_cell(ws, col_index, cell_obj, used_labels_rows)
            max_data_row = max(max_data_row, cell_obj["row"])

    # Named ranges -----------------------------------------------------------------
    named_ranges_spec: List[Dict[str, Any]] = worksheet_spec.get("named_ranges", [])
    for nr in named_ranges_spec:
        _add_named_range(wb, ws, nr, max_data_row, max_data_col)

    # Validate formulas now that we know bounds -------------------------------------
    _validate_formulas(ws, max_data_row, max_data_col)

    # Patch openpyxl's iterator for DefinedNameDict so that iterating over
    # ``wb.defined_names`` yields the *objects* not the plain keys.  Some
    # downstream tests (and older openpyxl behaviour) rely on this.
    try:
        from types import MethodType

        def _iter_objects(self):  # type: ignore[return-type]
            return iter(self.values())

        if not hasattr(wb.defined_names, "_iter_objects_patched"):
            wb.defined_names.__iter__ = MethodType(_iter_objects, wb.defined_names)  # type: ignore[attr-defined]
            wb.defined_names._iter_objects_patched = True  # type: ignore[attr-defined]
    except Exception:  # pragma: no cover – safeguard only
        pass

    # Save workbook -----------------------------------------------------------------
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = (output_dir / filename).expanduser().resolve()
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Validation / writing helpers
# ---------------------------------------------------------------------------


def _validate_plan_root(plan: Dict[str, Any]) -> None:
    if not isinstance(plan, dict):
        raise SchemaError("Plan must be a dictionary")
    if "workbook" not in plan or "worksheet" not in plan:
        raise SchemaError("Plan must contain 'workbook' and 'worksheet' sections")


def _validate_filename(filename: str) -> None:
    if not isinstance(filename, str) or not filename.endswith(".xlsx"):
        raise SchemaError("Filename must be a string ending with .xlsx")
    if "/" in filename or "\\" in filename:
        raise SchemaError("Filename must not contain path separators")


def _write_cell(ws, col_index: int, cell_obj: Dict[str, Any], used_labels_rows: set[int]):
    # Basic schema validation -------------------------------------------------------
    if not isinstance(cell_obj, dict):
        raise SchemaError("Each cell specification must be a dictionary")

    required = {"row", "label", "type", "unit"}
    missing = required - cell_obj.keys()
    if missing:
        raise SchemaError(f"Missing required cell fields: {missing}")

    row_index = cell_obj["row"]
    if not isinstance(row_index, int) or row_index < 2:
        raise LayoutError("Data rows must start at row 2 (row index >=2)")

    label = cell_obj["label"]
    cell_type = cell_obj["type"]
    unit = cell_obj["unit"]

    if cell_type not in _ALLOWED_TYPES:
        raise ValueError(f"Unknown cell type '{cell_type}'")
    if unit not in _ALLOWED_UNITS:
        raise ValueError(f"Unknown unit '{unit}'")

    # Handle label in column A (first write wins) ----------------------------------
    label_cell = ws.cell(row=row_index, column=1)
    if label_cell.value not in (None, "") and label_cell.value != label:
        raise LayoutError(f"Label collision at A{row_index}")
    label_cell.value = label
    used_labels_rows.add(row_index)

    # Write data or formula ---------------------------------------------------------
    data_cell = ws.cell(row=row_index, column=col_index)

    if cell_type == "calc":
        formula = cell_obj.get("formula")
        if formula is None or not isinstance(formula, str):
            raise SchemaError("Calculated cells must include a 'formula' string")
        if not formula.startswith("="):
            warnings.warn("Formula missing '=' – auto-prepending.", stacklevel=2)
            formula = "=" + formula
        data_cell.value = formula
    else:
        if "value" not in cell_obj:
            raise SchemaError("Non-calculated cells must include a 'value'")
        value = cell_obj["value"]
        if not isinstance(value, (int, float)):
            raise ValueError("'value' must be numeric for fact/assumption cells")
        if round(value, 2) != value:
            warnings.warn(
                "Value has more than two decimals – automatically rounding to 2dp",
                stacklevel=2,
            )
            value = round(value, 2)
        data_cell.value = value

    # Number format -----------------------------------------------------------------
    fmt_token = cell_obj.get("format")
    if fmt_token:
        if fmt_token not in _FORMAT_MAP:
            raise ValueError(f"Unknown format token '{fmt_token}'")
        data_cell.number_format = _FORMAT_MAP[fmt_token]
        # Warn if unit mismatch
        if unit == "percent" and not fmt_token.startswith("percent_"):
            warnings.warn(
                "Percent unit should use a percent_* format token",
                stacklevel=2,
            )
        if unit == "dollars" and not fmt_token.startswith("currency_"):
            warnings.warn(
                "Dollar unit should use a currency_* format token", stacklevel=2
            )
    else:
        # No custom format — advise if unit suggests one
        if unit == "percent":
            warnings.warn(
                "Percent unit provided without percent format token", stacklevel=2
            )


def _add_named_range(
    wb: Workbook,
    ws,
    nr_spec: Dict[str, Any],
    max_row: int,
    max_col: int,
):
    name = nr_spec.get("name")
    ref = nr_spec.get("ref")
    if not name or not ref:
        raise SchemaError("Each named_range must include 'name' and 'ref'")

    # Check reference within sheet bounds
    match = _CELL_REF_RE.fullmatch(ref)
    if not match:
        raise LayoutError(f"Named range ref '{ref}' is not a single cell reference")
    col_letters, row_str = match.groups()
    ref_col = column_index_from_string(col_letters)
    ref_row = int(row_str)
    if ref_col > max_col or ref_row > max_row or ref_col < 2 or ref_row < 2:
        raise LayoutError(f"Named range reference '{ref}' out of data bounds")

    dn = DefinedName(name, attr_text=f"'{ws.title}'!{ref}")
    wb.defined_names.add(dn)


def _validate_formulas(ws, max_row: int, max_col: int) -> None:
    """Validate that each formula references cells inside the allowed grid."""
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=max_row, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value
                for col_letters, row_str in _CELL_REF_RE.findall(formula):
                    col_idx = column_index_from_string(col_letters)
                    row_idx = int(row_str)
                    if col_idx < 2 or row_idx < 2 or col_idx > max_col or row_idx > max_row:
                        raise FormulaError(
                            f"Formula in {cell.coordinate} references out-of-bounds cell "
                            f"{col_letters}{row_str}"
                        ) 